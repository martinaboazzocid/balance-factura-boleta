"""
fetch_data.py - ZAS Chile - Factura vs Boleta
Fuentes:
  1. Odoo (JSON-RPC) - ventas Campanas Chile (fuente viva, se actualiza sola)
  2. ventas_historicas.xlsx - ventas 2021-2024 previas a Odoo (se sube una vez)
  3. pagos.xlsx - pagos realizados a talentos (se reemplaza cada semana)

Salida: data.json con ventas combinadas (Odoo + historico, deduplicado)
        y total pagado por talento.
"""

import json, os, http.cookiejar, urllib.request, unicodedata, re
from datetime import datetime, timezone
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    openpyxl = None

ODOO_URL  = "https://zas-talent.odoo.com"
ODOO_DB   = "zas-talent"
ODOO_USER = "martina.boazzo@zastalents.com"
ODOO_PASS = os.environ.get("ODOO_PASSWORD", "")

# Nombres fijos de los Excel en el repo
XLS_HISTORICO = "ventas_historicas.xlsx"   # se sube una vez
XLS_PAGOS     = "pagos.xlsx"               # se reemplaza cada semana

# Tipo de cambio USD->CLP: mensual (Odoo) + fallback anual (historico 2021-2023)
FX_CLP = {
    "2024-01":969,"2024-02":981,"2024-03":988,"2024-04":960,"2024-05":897,
    "2024-06":931,"2024-07":943,"2024-08":938,"2024-09":942,"2024-10":952,
    "2024-11":971,"2024-12":997,"2025-01":1010,"2025-02":998,"2025-03":990,
    "2025-04":969,"2025-05":944,"2025-06":951,"2025-07":953,"2025-08":945,
    "2025-09":955,"2025-10":963,"2025-11":971,"2025-12":981,"2026-01":988,
    "2026-02":970,"2026-03":960,"2026-04":945,"2026-05":935,"2026-06":942,
    "2026-07":948,"2026-08":945,"2026-09":950,"2026-10":955,"2026-11":960,
    "2026-12":965,
}
# Promedio anual BCCH (dolar observado) para ventas historicas
FX_YEAR = {"2021":780,"2022":870,"2023":840,"2024":945}

# -- Odoo --
_cj     = http.cookiejar.CookieJar()
_opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(_cj))

def odoo_call(endpoint, payload):
    data = json.dumps(payload).encode()
    req  = urllib.request.Request(
        ODOO_URL + endpoint, data=data,
        headers={"Content-Type": "application/json"}, method="POST"
    )
    with _opener.open(req, timeout=120) as r:
        return json.loads(r.read())

def odoo_auth():
    print("  Autenticando en Odoo...")
    r = odoo_call("/web/session/authenticate", {
        "jsonrpc":"2.0","method":"call","id":1,
        "params":{"db":ODOO_DB,"login":ODOO_USER,"password":ODOO_PASS}
    })
    uid = r["result"]["uid"]
    if not uid:
        raise Exception("Autenticacion fallida")
    print(f"  OK UID={uid}")

def search_read(model, domain, fields, batch=500):
    all_recs, offset = [], 0
    while True:
        r = odoo_call("/web/dataset/call_kw", {
            "jsonrpc":"2.0","method":"call","id":2,
            "params":{
                "model":model,"method":"search_read","args":[domain],
                "kwargs":{
                    "fields":fields,"limit":batch,"offset":offset,
                    "context":{"allowed_company_ids":[1,2,3,4,5,6]}
                }
            }
        })
        recs = r.get("result")
        if recs is None:
            raise Exception(f"Error {model}: {r.get('error')}")
        all_recs.extend(recs)
        print(f"    {model}: {len(all_recs)}...", end="\r")
        if len(recs) < batch:
            break
        offset += batch
    print(f"    {model}: {len(all_recs)} registros        ")
    return all_recs

# -- Utilidades --
def norm(s):
    if not s:
        return ""
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s

def mes_key(fecha):
    return fecha[:7] if fecha else "0000-00"

def to_clp(monto, moneda, fecha):
    moneda = (moneda or "CLP").upper().strip()
    if moneda == "CLP":
        return float(monto or 0)
    if moneda == "USD":
        rate = FX_CLP.get(mes_key(fecha), 950)
        return float(monto or 0) * rate
    return float(monto or 0)

def parse_talento(nombre):
    if not nombre:
        return None
    idx = nombre.find("(")
    raw = nombre[:idx].strip() if idx > 0 else nombre.strip()
    return norm(raw)

# -- Odoo utils --
def fetch_ordenes():
    print("\n  Descargando ordenes Campanas Chile...")
    orders = []
    for valor in ["Campa\u00f1as Chile", "Campanas Chile"]:
        orders = search_read(
            "sale.order",
            [
                ["x_studio_campaas_1", "=", valor],
                ["x_studio_factura_o_boleta", "in", ["Factura", "Boleta"]],
                ["state", "in", ["sale", "done"]]
            ],
            ["id","name","date_order","currency_id","amount_untaxed",
             "x_studio_factura_o_boleta","state"]
        )
        if orders:
            print(f"  Encontradas con: '{valor}'")
            break
    print(f"  {len(orders)} ordenes")
    return orders

def fetch_lineas(order_ids):
    if not order_ids:
        return []
    print("\n  Descargando lineas...")
    return search_read(
        "sale.order.line",
        [["order_id","in",order_ids],["state","in",["sale","done"]]],
        ["id","order_id","name","price_subtotal","currency_id"]
    )

# -- Procesamiento Odoo --
def procesar_odoo(orders, lines):
    ord_idx = {o["id"]: o for o in orders}

    general = defaultdict(lambda: {"f":0.0,"b":0.0,"usd":False})
    for o in orders:
        moneda  = (o.get("currency_id") or [None,"CLP"])[1] or "CLP"
        fecha   = (o.get("date_order") or "")[:10]
        mes     = mes_key(fecha)
        net_clp = to_clp(o.get("amount_untaxed",0), moneda, fecha)
        tipo    = o.get("x_studio_factura_o_boleta","")
        if moneda == "USD":
            general[mes]["usd"] = True
        if   tipo == "Factura": general[mes]["f"] += net_clp
        elif tipo == "Boleta":  general[mes]["b"] += net_clp

    talentos_mes = defaultdict(lambda: defaultdict(lambda: {"f":0.0,"b":0.0,"ordenes":[]}))
    for l in lines:
        nombre  = l.get("name") or ""
        talento = parse_talento(nombre)
        if not talento or len(talento) < 2:
            continue
        oid    = (l.get("order_id") or [None])[0]
        orden  = ord_idx.get(oid)
        if not orden:
            continue
        moneda  = (orden.get("currency_id") or [None,"CLP"])[1] or "CLP"
        fecha   = (orden.get("date_order") or "")[:10]
        mes     = mes_key(fecha)
        sub_clp = to_clp(l.get("price_subtotal",0), moneda, fecha)
        sub_orig = float(l.get("price_subtotal",0) or 0)
        tipo    = orden.get("x_studio_factura_o_boleta","")
        ord_name = orden.get("name","")

        if   tipo == "Factura": talentos_mes[talento][mes]["f"] += sub_clp
        elif tipo == "Boleta":  talentos_mes[talento][mes]["b"] += sub_clp

        if tipo in ("Factura", "Boleta"):
            talentos_mes[talento][mes]["ordenes"].append({
                "nombre": nombre, "orden": ord_name, "tipo": tipo,
                "monto_clp": round(sub_clp), "monto_orig": round(sub_orig),
                "moneda": moneda, "fecha": fecha, "fuente": "odoo",
            })
    return general, talentos_mes

# -- Ventas historicas (Excel) --
def tipo_documento(ag):
    a = (ag or "").strip().upper()
    if a == "F": return "Factura"
    if a == "B": return "Boleta"
    if a == "REBATE": return None
    if a in ("PAYONEER USD", "PAYPAL", "BTC"): return "Factura"
    return None

def cargar_historico(general, talentos_mes, odoo_keys):
    if openpyxl is None or not os.path.exists(XLS_HISTORICO):
        print(f"  (sin {XLS_HISTORICO}, se omiten ventas historicas)")
        return 0, 0
    print(f"\n  Cargando ventas historicas desde {XLS_HISTORICO}...")
    wb = openpyxl.load_workbook(XLS_HISTORICO, data_only=True)
    ws = wb.worksheets[0]
    added = dup = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        tipo = tipo_documento(r[32] if len(r) > 32 else None)
        if not tipo:
            continue
        tal = norm(r[7] if len(r) > 7 else "")
        if not tal or len(tal) < 2 or tal in ("NA", "N/A", "-", "X"):
            continue
        y = r[59] if len(r) > 59 else None
        m = r[60] if len(r) > 60 else None
        if not y or not m:
            continue
        mes = f"{int(y):04d}-{int(m):02d}"
        try:
            monto = float(r[6] or 0)
        except (TypeError, ValueError):
            continue
        moneda = (r[5] or "CLP").strip().upper()
        monto_orig = monto
        if moneda == "USD":
            monto *= FX_YEAR.get(str(int(y)), 850)

        key = (tal, mes, round(monto))
        if key in odoo_keys:
            dup += 1
            continue

        if tipo == "Factura":
            talentos_mes[tal][mes]["f"] += monto
            general[mes]["f"] += monto
        else:
            talentos_mes[tal][mes]["b"] += monto
            general[mes]["b"] += monto
        if moneda == "USD":
            general[mes]["usd"] = True

        talentos_mes[tal][mes]["ordenes"].append({
            "nombre": r[4] or tal, "orden": r[1] or "", "tipo": tipo,
            "monto_clp": round(monto), "monto_orig": round(monto_orig),
            "moneda": moneda, "fecha": mes + "-01", "fuente": "historico",
        })
        added += 1
    print(f"  Historico: {added} ventas agregadas, {dup} duplicados descartados")
    return added, dup

# -- Pagos a talentos (Excel) --
def cargar_pagos():
    if openpyxl is None or not os.path.exists(XLS_PAGOS):
        print(f"  (sin {XLS_PAGOS}, sin datos de pagos)")
        return {}
    print(f"\n  Cargando pagos desde {XLS_PAGOS}...")
    wb = openpyxl.load_workbook(XLS_PAGOS, data_only=True)
    ws = wb.worksheets[0]
    pagos = defaultdict(float)
    conteo = defaultdict(int)
    for r in ws.iter_rows(min_row=2, values_only=True):
        bu = norm(r[15] if len(r) > 15 else "")
        if bu != "ZAS CHILE":
            continue
        tal = norm(r[5] if len(r) > 5 else "")
        if not tal or len(tal) < 2:
            continue
        try:
            monto = float(r[14] or 0)
        except (TypeError, ValueError):
            continue
        pagos[tal] += monto
        conteo[tal] += 1
    out = {t: {"total": round(v), "n": conteo[t]} for t, v in pagos.items()}
    print(f"  Pagos: {len(out)} talentos con pagos ZAS CHILE")
    return out

# -- Main --
def main():
    print("=== fetch_data.py - ZAS Chile ===")
    print(f"Inicio: {datetime.now(timezone.utc).isoformat()}")

    if not ODOO_PASS:
        raise RuntimeError("Falta ODOO_PASSWORD")

    odoo_auth()
    orders = fetch_ordenes()
    order_ids = [o["id"] for o in orders]
    lines = fetch_lineas(order_ids)

    print(f"\n  Procesando {len(orders)} ordenes Odoo, {len(lines)} lineas...")
    general, talentos_mes = procesar_odoo(orders, lines)

    odoo_keys = set()
    for t, meses in talentos_mes.items():
        for m, v in meses.items():
            for o in v["ordenes"]:
                odoo_keys.add((t, m, round(o["monto_clp"])))

    cargar_historico(general, talentos_mes, odoo_keys)
    pagos = cargar_pagos()

    talentos_out = {}
    for t, meses in talentos_mes.items():
        talentos_out[t] = {}
        for m, vals in meses.items():
            talentos_out[t][m] = {
                "f": vals["f"], "b": vals["b"], "ordenes": vals["ordenes"],
            }

    output = {
        "updated_at":    datetime.now(timezone.utc).isoformat(),
        "total_ordenes": len(orders),
        "general":       dict(general),
        "talentos":      talentos_out,
        "pagos":         pagos,
    }

    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n  OK data.json generado")
    print(f"  Meses: {min(general)} -> {max(general)}")
    print(f"  Talentos con ventas: {len(talentos_out)} | con pagos: {len(pagos)}")

if __name__ == "__main__":
    main()
